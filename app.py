from flask import Flask,render_template,redirect,request,url_for,Response
import mysql.connector
import cv2,os,csv
from PIL import Image
import time
import numpy as np
from PIL import Image, ImageOps  # Install pillow instead of PIL
from imutils import paths
import face_recognition
import pickle
import pandas as pd
from openpyxl import workbook,load_workbook
from datetime import datetime
import smtplib




app = Flask(__name__)



#home display
@app.route("/")
def home():
    return render_template('login.html')

#login page display
@app.route("/loginsh")
def loginsh():
    return render_template("login.html")

#teacher module display
@app.route("/teach")
def teach():
    return render_template('teachermodule.html')

#attendance edit page display
@app.route("/update")
def update():
    return render_template('edit.html')



#admin page display
@app.route('/adminsh')
def adminsh():
    return render_template('admin.html')

#admini page display
@app.route("/admini")
def admini():
    return render_template("admini.html")

#teacher value entry page display
@app.route("/techent")
def techent():
    return render_template("teacherentry.html")


#student value entry page display
@app.route("/studentry")
def studentry():
    return render_template("studentry.html")

#facecapturing page display
@app.route("/facecap")
def facecap():
    return render_template('facecap.html')



#redirection when login submit
@app.route("/login",methods=['POST'])
def login():
    username=request.form['username']
    password=request.form['password']
    mydb=mysql.connector.connect(host="localhost",user="root",password="root",database="adm")
    mycursor = mydb.cursor()
    sql = "SELECT teachid,password FROM teacher WHERE teachid = %s AND password = %s"
    val = (username, password)
    mycursor.execute(sql, val)
    result = mycursor.fetchone()
    if result:
        return render_template('teachermodule.html')
    else:
         error_message = "Invalid username and password ! enter correct credential"
         return render_template("login.html", error=error_message)



#attendance view for teacher module
@app.route("/att")
def att():
    book=load_workbook('.venv/attendance.xlsx')
    sheet=book.active
    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
     data.append((str(row[0]),) + tuple(str(value) for value in row[2:]))
    return render_template('attendanceteach.html',data=data)


#searching attendnace based on date for teacher module
@app.route('/searchteach', methods=['POST'])
def searchteach():
    search_query = request.form['search']     
    workbook = load_workbook('.venv/permanentattendance.xlsx')
    sheet = workbook.active
    data = []
    for row in sheet.iter_rows(values_only=True):
        if row[1] == search_query:
            data.append((str(row[0]),) + tuple(str(value) for value in row[2:]))
    return render_template('attendanceteach.html', data=data)


#editing attendance code
@app.route('/update_attendance', methods=['POST'])
def update_attendance():
    roll_no = request.form["rollnumber"]
    period = int(request.form["period"])
    status = request.form["status"]
    workbook = load_workbook('.venv/attendance.xlsx')
    sheet = workbook.active
    column_index = period + 2  
    for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if row[0] == roll_no:
            # Update the status in the corresponding cell
            cell = sheet.cell(row=row_number, column=column_index)
            cell.value = status
            break  # Exit the loop once the cell is found and updated

    workbook.save('.venv/attendance.xlsx')
    update= 'Attendance updated successfully'
    return render_template('edit.html',update=update)


# send email 
@app.route("/emailsend")
def emailsend():
    attendance_df = pd.read_excel('.venv/attendance.xlsx')
    student_details_df = pd.read_excel('.venv/student_details.xlsx')

    # SMTP server details
    smtp_server = 'smtp.gmail.com'
    smtp_port = 587  # Port for TLS

    sender_email = 'spot8776@gmail.com'
    sender_password = 'eowrcuzsvzbxklrc'

    # Iterate over each student
    for _, row in student_details_df.iterrows():
        roll_number = row['Roll Number']
        student_email = row['Email']

        # Filter attendance records for the specific student
        student_attendance = attendance_df.loc[attendance_df['Rollno'] == roll_number]

        # Compose the email
        subject = 'Attendance Report'
        message = f'Dear Student,\n\nHere is your attendance report:\n\n{student_attendance.to_string(index=False)}\n\nSincerely,\nAkash Global college of Management and Science'

        # Send the email with the attendance report
        with smtplib.SMTP(smtp_server, smtp_port) as server:
            server.starttls()
            server.login(sender_email, sender_password)
            server.sendmail(sender_email, student_email, f'Subject: {subject}\n\n{message}')

    sent = "Emails sent successfully"
    return render_template("teachermodule.html", sent=sent)

#redirection when admin submit
@app.route("/admin",methods=['POST'])
def admin():
    username=request.form['username']
    password=request.form['password']
    mydb=mysql.connector.connect(host="localhost",user="root",password="root",database="adm")
    mycursor = mydb.cursor()
    sql = "SELECT * FROM ad WHERE username = %s AND password = %s"
    val = (username, password)
    mycursor.execute(sql, val)
    result = mycursor.fetchone()
    if result:
        return redirect("admini")
    else:
        error_message = "Invalid username and password ! enter correct credential"
        return render_template("admin.html", error=error_message)

        
    


#teacher value storing to database
@app.route("/teachervalue", methods=["POST"])
def teachervalue():
    id = request.form["id"]
    name = request.form["name"]
    password = request.form["password"]
    email = request.form["email"]
    subject = request.form["subject"]
    
    conn = mysql.connector.connect(host="localhost", user="root", password="root", database="adm")
    cursor = conn.cursor()
    
    # Check if the ID already exists in the database
    query = "SELECT COUNT(*) FROM teacher WHERE teachid = %s"
    cursor.execute(query, (id,))
    result = cursor.fetchone()
    
    if result[0] > 0:
        # ID already exists, show an error message
        error_message = "Teacher ID already exists. Please choose a different ID."
        return render_template("teacherentry.html", error=error_message)
    else:
        # ID is unique, insert the teacher values into the database
        sql = "INSERT INTO teacher (teachid, name, password, subject, email) VALUES (%s, %s, %s, %s, %s)"
        values = (id, name, password, subject, email)
        cursor.execute(sql, values)
        conn.commit()
        cursor.close()
        success_message = "Teacher details saved successfully."
        return render_template("teacherentry.html", success=success_message)




#teacher details showing     
@app.route("/techview")
def techview():
    conn=mysql.connector.connect(host="localhost",user="root",password="root",database="adm")
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM teacher")
    data = cursor.fetchall()
    return render_template("teach1.html",data=data)





#search button code for teacher value
@app.route('/search', methods=['POST'])
def search():
    name = request.form['search']
    conn=mysql.connector.connect(host="localhost",user="root",password="root",database="adm")
    cursor = conn.cursor()
    query="SELECT * FROM teacher where name like %s"
    cursor.execute(query,(name,))
    data = cursor.fetchall()
    return render_template("teach1.html",data=data)





#teacher row delete
@app.route('/delete_row/<string:id>', methods=['POST'])
def delete_row(id):
    # Delete the row from the database
    connection = mysql.connector.connect(
        host="localhost",user="root",password="root",database="adm"
    )
    cursor = connection.cursor()
    cursor.execute("DELETE FROM teacher WHERE teachid= %s", (id,))
    connection.commit()
    connection.close()
    return redirect(url_for('techview'))




#student value storing to database
@app.route("/studententry", methods=["POST"])
def studententry():
    name = request.form["name"]
    class1 = request.form["class"]
    rollno = request.form["rollno"]
    email = request.form["email"]
    image = "no"
    
    # Check if the roll number already exists
    existing_rollnos = pd.read_excel('.venv/student_details.xlsx', usecols=["Roll Number"])["Roll Number"].tolist()
    if rollno in existing_rollnos:
        error_message = "Roll number already exists."
        return render_template("studentry.html", error=error_message)
    # Insert into Excel sheet
    data = {
        'Name': [name],
        'Class': [class1],
        'Roll Number': [rollno],
        'Email': [email],
        'Image': [image]
    }
    df = pd.DataFrame(data)
    filename = '.venv/student_details.xlsx'
    try:
        existing_df = pd.read_excel(filename)
        df = pd.concat([existing_df, df], ignore_index=True)
    except FileNotFoundError:
        pass  # The file doesn't exist, so we'll create a new one
    df.to_excel(filename, index=False)
    
    success_message = "Student details saved successfully."
    return render_template("studentry.html", success=success_message)





#student details showing     
@app.route("/studentview")
def studentview():
    book=load_workbook('.venv/student_details.xlsx')
    sheet=book.active
    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        data.append(row)
    return render_template("studentdisp.html", data=data)





#deleting student row
@app.route('/delete_rows/<string:id>', methods=['POST'])
def delete_rows(id):
    book = load_workbook('.venv/student_details.xlsx')
    sheet = book.active
    rows_to_delete = []
    rows = list(sheet.iter_rows(min_row=2))
    for row in reversed(rows):
        if row[2].value == id:
            rows_to_delete.append(row[2].row)

    for row_index in rows_to_delete:
        sheet.delete_rows(row_index)

    book.save('.venv/student_details.xlsx')
    return redirect(url_for('studentview'))





#search button code for std value
@app.route('/searchstd', methods=['POST'])
def searchstd():
    search_query = request.form.get('search')  # Get the search query from the form
    
    # Load the Excel file
    
    workbook = load_workbook('.venv/student_details.xlsx')
    sheet = workbook.active
    
    # Search for the required field
    data = []
    for row in sheet.iter_rows(values_only=True):
        if row[0] == search_query:
            data.append(row)
    
    return render_template('studentdisp.html', data=data)




#sample  pic taking 
@app.route('/take_images', methods=['POST'])
def take_images():
    rollno = request.form['roll_number']
    base_dir = ".venv/static/studentpic"  # Base directory to store images
    rollno_dir = os.path.join(base_dir, rollno)  # Directory specific to the roll number
    # Create the roll number directory if it doesn't exist
    if not os.path.isdir(rollno_dir):
        os.makedirs(rollno_dir)
    # Initialize the camera and face detector
    cam = cv2.VideoCapture(0)
    detector = cv2.CascadeClassifier('.venv/haarcascade_frontalface_default.xml')
    sampleNum = 0
    while True:
        ret, img = cam.read()
        if not ret:
            return "Failed to capture image from camera"
        faces = detector.detectMultiScale(img, 1.3, 5, minSize=(30, 30), flags=cv2.CASCADE_SCALE_IMAGE)
        for (x, y, w, h) in faces: 

            x -= 20
            y -= 20
            w += 40
            h += 40
            cv2.rectangle(img, (x, y), (x+w, y+h), (10, 159, 255), 2)
            sampleNum += 1
            image_path = os.path.join(rollno_dir, f"{rollno}_{sampleNum}.jpg")
            cv2.imwrite(image_path, img[y:y+h, x:x+w])
            cv2.imshow('frame', img)
        if cv2.waitKey(100) & 0xFF == ord('q') or sampleNum > 100:
            break
    cam.release()
    cv2.destroyAllWindows()
    
    #updating rollno to csv file
    csv_path = ".venv/studentdetails/studentdetails.csv"
    header = ["Roll No"]
    row = [rollno]
    
    if os.path.isfile(csv_path):
        with open(csv_path, 'a+', newline='') as csvFile:
            writer = csv.writer(csvFile)
            writer.writerow(row)
    else:
        with open(csv_path, 'w', newline='') as csvFile:
            writer = csv.writer(csvFile)
            writer.writerow(header)
            writer.writerow(row)

    #updateing yes to image for student display       
    filename = '.venv/student_details.xlsx'
    df = pd.read_excel(filename)
    df.loc[df['Roll Number'] == rollno, 'Image'] = 'yes'
    df.to_excel(filename, index=False)

    #adding rollnumber to attendance excel
    filename2='.venv/attendance.xlsx'
    book = load_workbook(filename2)
    sheet = book.active
    empty_row = sheet.max_row + 1
    sheet[f'A{empty_row}'] = rollno
    book.save(filename2)
    msg="Image taken successfully for roll number  "+ rollno
    return render_template('facecap.html',msg=msg)




#sample picture encoding
@app.route('/encode_faces')
def encode_faces():
    # Specify the main directory path
    main_directory = '.venv/static/studentpic'
    # Get paths of each file in the subfolders
    imagePaths = list(paths.list_images(main_directory))
    # Rest of the code remains the same
    knownEncodings = []
    knownNames = []
    # Loop over the image paths
    for (i, imagePath) in enumerate(imagePaths):
        # Extract the person name from the image path
        name = imagePath.split(os.path.sep)[-2]
        # Load the input image and convert it from BGR (OpenCV ordering)
        # to dlib ordering (RGB)
        image = cv2.imread(imagePath)
        rgb = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)
        # Use face_recognition to locate faces
        boxes = face_recognition.face_locations(rgb, model='hog')
        # Compute the facial embeddings for the face
        encodings = face_recognition.face_encodings(rgb, boxes)
        # Loop over the encodings
        for encoding in encodings:
            knownEncodings.append(encoding)
            knownNames.append(name)
    data = {"encodings": knownEncodings, "names": knownNames}
    with open("face_enc.pkl", "wb") as f:
        pickle.dump(data, f)
    msg="face encoding saved"
    return render_template('admini.html',msg=msg)




#take attendance
@app.route('/video_feed')
def video_feed():
    cascPathface = ".venv/haarcascade_frontalface_default.xml"
    faceCascade = cv2.CascadeClassifier(cascPathface)
    data = pickle.loads(open('face_enc.pkl', "rb").read())

    def generate_frames():
        video_capture = cv2.VideoCapture(0)
        while True:
            ret, frame = video_capture.read()
            gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
            faces = faceCascade.detectMultiScale(gray,
                                                 scaleFactor=1.1,
                                                 minNeighbors=5,
                                                 minSize=(60, 60),
                                                 flags=cv2.CASCADE_SCALE_IMAGE)

            rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
            encodings = face_recognition.face_encodings(rgb)
            names = []
            for encoding in encodings:
                face_distances = face_recognition.face_distance(data["encodings"], encoding)
                matches = face_recognition.compare_faces(data["encodings"], encoding, tolerance=0.6)
                name = "Unknown"

                if True in matches:
                    # Find the index with the smallest distance
                    best_match_index = np.argmin(face_distances)
                    if matches[best_match_index]:
                        name = data["names"][best_match_index]
                        if face_distances[best_match_index] > 0.6:
                            name = "Unknown"
                        #attendance marking calling mark function
                        mark(name)    
                names.append(name)

            for ((x, y, w, h), name) in zip(faces, names):
                cv2.rectangle(frame, (x, y), (x + w, y + h), (0, 255, 0), 2)
                cv2.putText(frame, name, (x, y), cv2.FONT_HERSHEY_SIMPLEX, 0.75, (0, 255, 0), 2)
            ret, buffer = cv2.imencode('.jpg', frame)
            frame = buffer.tobytes()
            yield (b'--frame\r\n'
                   b'Content-Type: image/jpeg\r\n\r\n' + frame + b'\r\n')

        video_capture.release()
        cv2.destroyAllWindows()

    return Response(generate_frames(), mimetype='multipart/x-mixed-replace; boundary=frame')


#attendance marking code
def mark(name):
    workbook = load_workbook('.venv/attendance.xlsx')
    sheet = workbook.active

    # mark every empty cell in attendance excel as absent
    for row in sheet.iter_rows():
      for cell in row:
        if cell.value is None or cell.value == "":
            cell.value = "absent"
    workbook.save('.venv/attendance.xlsx')

    now = datetime.now()
    current_time = now.strftime("%H:%M:%S")
    curr_date = now.strftime('%d-%m-%y')
    wb = load_workbook('.venv/attendance.xlsx')
    ws = wb.active
    _9_clock = '09:00:00'
    _915_clock = '09:15:00'
    _10_clock = '10:00:00'
    _1015_clock = '10:30:00'
    _11_clock = '11:00:00'
    _1115_clock = '11:15:00'
    _12_clock = '12:00:00'
    _1215_clock = '12:15:00'
    _2_clock = '14:00:00'
    _215_clock = '14:15:00'
    _3_clock = '15:00:00'
    _315_clock = '15:15:00'
    _4_clock = '16:00:00'
    _415_clock = '16:15:00'
    
    column_a = ws['A']
    column_b = ws['B']
    for i, cell in enumerate(column_b, start=2):
        if i > 2:  # Exclude the first row
            cell.value = curr_date
    rollno = name
    i = 0
    for cell in column_a:
        i += 1
        if rollno in cell.value:
            if _9_clock < current_time < _915_clock:
                if ws.cell(row=i, column=3).value == 'present':
                    pass
                else:
                    ws.cell(row=i, column=3).value = 'present'
            elif _10_clock < current_time < _1015_clock:
                if ws.cell(row=i, column=4).value == 'present':
                    pass
                else:
                    ws.cell(row=i, column=4).value = 'present'
            elif _11_clock < current_time < _1115_clock:
                if ws.cell(row=i, column=5).value == 'present':
                    pass
                else:
                    ws.cell(row=i, column=5).value = 'present'
            elif _12_clock < current_time < _1215_clock:
                if ws.cell(row=i, column=6).value == 'present':
                    pass
                else:
                    ws.cell(row=i, column=6).value = 'present'
            elif _2_clock < current_time < _215_clock:
                if ws.cell(row=i, column=7).value == 'present':
                    pass
                else:
                    ws.cell(row=i, column=7).value = 'present'
            elif _3_clock < current_time < _315_clock:
                if ws.cell(row=i, column=8).value == 'present':
                    pass
                else:
                    ws.cell(row=i, column=8).value = 'present'
            elif _4_clock < current_time < _415_clock:
                if ws.cell(row=i, column=9).value == 'present':
                    pass
                else:
                    ws.cell(row=i, column=9).value = 'present'
    wb.save('.venv/attendance.xlsx')



#attendance display from attendance excel
@app.route("/attenddis")
def attenddis():
    book=load_workbook('.venv/attendance.xlsx')
    sheet=book.active
    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
     data.append((str(row[0]),) + tuple(str(value) for value in row[2:]))  # Convert tuple elements to strings

    return render_template("attendancesheet.html", data=data)



#code for save button clear attendance sheet except first col and row append attendance of day to permanentattendance   
@app.route('/permanentsave')
def permanentsave():
    current_time = datetime.now().time()
    if current_time.hour < 17:
        permansaveerror="Data saving is only allowed after 5 PM."
        return render_template("admini.html", permansaveerror=permansaveerror)
    workbook = load_workbook('.venv/attendance.xlsx')
    sheet = workbook.active

    destination=load_workbook('.venv/permanentattendance.xlsx')
    dessheet=destination.active
    for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
      if row_index != 1:  # Skip the first row
        dessheet.append(row)
    
    #deleting values from attendance excel
    max_row = sheet.max_row
    max_column = sheet.max_column
    for row in sheet.iter_rows(min_row=2, max_row=max_row, min_col=2, max_col=max_column):
     for cell in row:
        cell.value = None
    destination.save('.venv/permanentattendance.xlsx')
    workbook.save('.venv/attendance.xlsx')
    permansave="Attendance saved"
    return render_template("admini.html", permansave=permansave)
    



#search attendnace based on date
@app.route('/searchatten', methods=['POST'])
def searchatten():
    search_query = request.form['search'] # Get the search query from the form
    
    # Load the Excel file
    
    workbook = load_workbook('.venv/permanentattendance.xlsx')
    sheet = workbook.active
    
    # Search for the required field
    data = []
    for row in sheet.iter_rows(values_only=True):
        if row[1] == search_query:
            data.append((str(row[0]),) + tuple(str(value) for value in row[2:]))
    
    return render_template('attendancesheet.html', data=data)



#generating monthly report and storing
@app.route('/generate_report')
def generate_report():
    
    df = pd.read_excel('.venv/permanentattendance.xlsx')
    df['date'] = pd.to_datetime(df['date'], format='%d-%m-%y')
    # Calculate the total number of days in the month
    total_days = len(df['date'].unique())
    total_classes = len(df.columns) - 2  
    attendance_columns = df.columns[2:]
    df[attendance_columns] = df[attendance_columns].astype(str)
    df['Total Attendance'] = df[attendance_columns].apply(lambda row: row.str.count('present').sum(), axis=1)
    df['Attendance Percentage'] = (df['Total Attendance'] / total_classes) * 100
    attendance_report = df.groupby('Rollno').agg(
        Total_Classes=('Total Attendance', lambda x: total_days * 8),
        Total_Attendance=('Total Attendance', 'sum'),
        Attendance_Percentage=('Attendance Percentage', lambda x: round(x.mean(), 2))

    ).reset_index()
    report_file = '.venv/monthly_report.xlsx'
    attendance_report.to_excel(report_file, index=False)

    gen = 'Attendance report generated.'
    return render_template("teachermodule.html", gen=gen)


#viewing the report
@app.route("/viewreport")
def viewreport():
    book=load_workbook('.venv/monthly_report.xlsx')
    sheet=book.active
    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
     data.append((str(row[0]), str(row[1])) + tuple(str(value) for value in row[2:]))


    return render_template("viewreport.html", data=data)



#sending the report
@app.route("/sendreport")
def sendreport():
    attendance_df = pd.read_excel('.venv/monthly_report.xlsx')
    student_details_df = pd.read_excel('.venv/student_details.xlsx')

    # SMTP server details
    smtp_server = 'smtp.gmail.com'
    smtp_port = 587  # Port for TLS

    sender_email = 'spot8776@gmail.com'
    sender_password = 'eowrcuzsvzbxklrc'

    # Iterate over each student
    for _, row in student_details_df.iterrows():
        roll_number = row['Roll Number']
        student_email = row['Email']

        # Filter attendance records for the specific student
        student_attendance = attendance_df.loc[attendance_df['Rollno'] == roll_number]

        # Compose the email
        subject = 'Attendance Report'
        message = f'Dear Student,\n\nHere is your montly attendance report:\n\n{student_attendance.to_string(index=False)}\n\nSincerely,\nAGCMS college'

        # Send the email with the attendance report
        with smtplib.SMTP(smtp_server, smtp_port) as server:
            server.starttls()
            server.login(sender_email, sender_password)
            server.sendmail(sender_email, student_email, f'Subject: {subject}\n\n{message}')

    sent = "Emails sent successfully"
    return render_template("viewreport.html", sent=sent)



if __name__ == "__main__":
    app.run(debug=True)


